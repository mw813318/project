import logging
import asyncio
import sqlite3
import os
from datetime import datetime
from aiogram import Bot, Dispatcher, Router, F
from aiogram.types import Message, FSInputFile
from aiogram.filters import CommandStart, Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
import openpyxl
from openpyxl.utils import get_column_letter

API_TOKEN = "7853064330:AAF4Yjnox6_psclnspp4_j1c0gwCR5NcEvk"
GROUP_CHAT_ID = -1002558497144  # Замените на ваш реальный ID

logging.basicConfig(level=logging.INFO)

bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
router = Router()

DB_PATH = 'test_requests.db'

class Form(StatesGroup):
    supplier = State()
    amount = State()
    agent_name = State()
    agent_phone = State()
    delivery_date = State()
    description = State()
    admin_name = State()

def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT,
                supplier TEXT,
                amount REAL,
                agent_name TEXT,
                agent_phone TEXT,
                delivery_date TEXT,
                description TEXT,
                admin_name TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

def generate_excel_by_date(date_str):
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT supplier, amount, agent_name, agent_phone, delivery_date, description, admin_name, username 
            FROM requests WHERE delivery_date = ?
        """, (date_str,))
        rows = cur.fetchall()

    if not rows:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    ws.append(['Поставщик','Сумма','Имя агента','Номер','Дата поставки','Описание','Админ','От кого'])
    total = 0

    for row in rows:
        amount = float(row[1])
        ws.append([
            row[0], amount, row[2], row[3],
            row[4], row[5], row[6], row[7]
        ])
        total += amount

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#,##0'

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

    ws.append([])
    total_formatted = "{:,.0f}".format(total).replace(",", ".")
    ws.append(['', '💰 Общая сумма:', total_formatted])

    filename = f"Заявки_{date_str}.xlsx"
    wb.save(filename)
    return filename

@router.message(CommandStart())
async def start(message: Message):
    await message.answer("Привет! Это тестовый бот.\n\nДоступные команды:\n/заявка — создать заявку\n/экспорт — экспортировать в Excel")

@router.message(Command("заявка"))
async def start_form(message: Message, state: FSMContext):
    await state.set_state(Form.supplier)
    await message.answer("Введи поставщика:")

@router.message(Form.supplier)
async def step_supplier(message: Message, state: FSMContext):
    await state.update_data(supplier=message.text)
    await state.set_state(Form.amount)
    await message.answer("Введи сумму:")

@router.message(Form.amount)
async def step_amount(message: Message, state: FSMContext):
    try:
        amount = float(message.text.replace(',', '.').replace(' ', ''))
    except:
        return await message.answer("Некорректная сумма. Введи число.")
    await state.update_data(amount=amount)
    await state.set_state(Form.agent_name)
    await message.answer("Имя агента:")

@router.message(Form.agent_name)
async def step_agent_name(message: Message, state: FSMContext):
    await state.update_data(agent_name=message.text)
    await state.set_state(Form.agent_phone)
    await message.answer("Номер агента:")

@router.message(Form.agent_phone)
async def step_agent_phone(message: Message, state: FSMContext):
    await state.update_data(agent_phone=message.text)
    await state.set_state(Form.delivery_date)
    await message.answer("Дата поставки (дд.мм.гггг):")

@router.message(Form.delivery_date)
async def step_delivery_date(message: Message, state: FSMContext):
    try:
        date_obj = datetime.strptime(message.text, "%d.%m.%Y").date()
        await state.update_data(delivery_date=date_obj.strftime("%Y-%m-%d"))
        await state.set_state(Form.description)
        await message.answer("Описание заявки:")
    except:
        await message.answer("Неверный формат даты. Введи ДД.ММ.ГГГГ")

@router.message(Form.description)
async def step_description(message: Message, state: FSMContext):
    await state.update_data(description=message.text)
    await state.set_state(Form.admin_name)
    await message.answer("Имя админа:")

@router.message(Form.admin_name)
async def step_admin_name(message: Message, state: FSMContext):
    await state.update_data(admin_name=message.text)
    data = await state.get_data()
    user = message.from_user

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO requests (user_id, username, supplier, amount, agent_name, agent_phone, delivery_date, description, admin_name)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user.id, user.full_name, data['supplier'], data['amount'],
            data['agent_name'], data['agent_phone'], data['delivery_date'], data['description'], data['admin_name']
        ))

    await state.clear()
    await message.answer("✅ Заявка сохранена.")

    # Отправка в группу
    await bot.send_message(GROUP_CHAT_ID,
        f"📦 Новая заявка от {user.full_name}:\n\n"
        f"Поставщик: {data['supplier']}\n"
        f"Сумма: {data['amount']}\n"
        f"Агент: {data['agent_name']}\n"
        f"Номер: {data['agent_phone']}\n"
        f"Дата поставки: {data['delivery_date']}\n"
        f"Описание: {data['description']}\n"
        f"Админ: {data['admin_name']}"
    )

@router.message(Command("экспорт"))
async def export_requests(message: Message):
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = generate_excel_by_date(date_str)
    if filename:
        await message.answer_document(FSInputFile(filename))
        os.remove(filename)
    else:
        await message.answer("Нет заявок для экспорта.")

async def main():
    init_db()
    dp.include_router(router)
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())
