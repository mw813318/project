import logging
import asyncio
import os
import sqlite3
from datetime import datetime, timedelta
from collections import defaultdict
from dotenv import load_dotenv

from aiogram import Bot, Dispatcher, Router, F
from aiogram.types import Message, FSInputFile
from aiogram.filters import CommandStart, Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
import openpyxl
from openpyxl.utils import get_column_letter

load_dotenv()

API_TOKEN = os.getenv("BOT_TOKEN")
GROUP_CHAT_ID = int(os.getenv("GROUP_CHAT_ID"))

logging.basicConfig(level=logging.INFO)

bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
router = Router()

DB_PATH = 'requests.db'

class Form(StatesGroup):
    supplier = State()
    amount = State()
    description = State()
    agent_name = State()
    agent_phone = State()
    delivery_date = State()
    admin_name = State()

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
   
    # Создаем таблицу, если ее нет
    cur.execute("""
        CREATE TABLE IF NOT EXISTS requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT,
            supplier TEXT,
            amount REAL,
            agent_name TEXT,
            agent_phone TEXT,
            delivery_date TEXT,
            admin_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
   
    # Проверяем наличие колонки description и добавляем если ее нет
    cur.execute("PRAGMA table_info(requests)")
    columns = [column[1] for column in cur.fetchall()]
    if 'description' not in columns:
        cur.execute("ALTER TABLE requests ADD COLUMN description TEXT")
   
    conn.commit()
    conn.close()

def generate_excel_by_date(date_str):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        SELECT supplier, amount, description, agent_name, agent_phone, delivery_date, admin_name, username
        FROM requests WHERE delivery_date = ?
    """, (date_str,))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    ws.append(['Поставщик','Сумма','Описание','Имя агента','Номер','Дата поставки','Админ','От кого'])
    total = 0

    for row in rows:
        amount = float(row[1])
        excel_row = [
            row[0],          # supplier
            amount,          # real number
            row[2],          # description
            row[3],          # agent_name
            row[4],          # agent_phone
            row[5],          # delivery_date
            row[6],          # admin_name
            row[7]           # username
        ]
        ws.append(excel_row)
        total += amount

    # Применяем числовой формат к колонке "Сумма"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#,##0'
   
    # Автоширина
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

    # Итог
    ws.append([])
    total_formatted = "{:,.0f}".format(total).replace(",", ".")
    ws.append(['', '💰 Общая сумма:', total_formatted])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

    filename = f"Заявки_{date_str}.xlsx"
    wb.save(filename)
    return filename

@router.message(CommandStart())
async def start(message: Message):
    await message.answer("Привет! Я бот для учета заявок. Используй команду /заявка чтобы создать новую заявку.")

@router.message(Command("поставки"))
async def show_deliveries(message: Message):
    parts = message.text.split()
    if len(parts) > 1:
        try:
            date_str = datetime.strptime(parts[1], "%d.%m.%Y").strftime("%Y-%m-%d")
        except ValueError:
            return await message.answer("❌ Неверный формат даты. Используй: /поставки дд.мм.гггг")
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        SELECT supplier, amount, description FROM requests
        WHERE delivery_date = ?
    """, (date_str,))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return await message.answer(f"📭 Нет поставок на {date_str}.")

    total = sum(float(row[1]) for row in rows)
    total_formatted = "{:,.0f}".format(total).replace(",", ".")
   
    text = f"📦 Поставки на {datetime.strptime(date_str, '%Y-%m-%d').strftime('%d.%m.%Y')}:\n"
    for i, row in enumerate(rows, 1):
        amount_formatted = "{:,.0f}".format(float(row[1])).replace(",", ".")
        text += f"\n{i}) Поставщик: {row[0]}\nСумма: {amount_formatted}"
        if row[2]:  # Добавляем описание, если оно есть
            text += f"\nОписание: {row[2]}"
        text += "\n"  # Добавляем дополнительный пробел между заявками

    text += f"\n💰 Общая сумма поставок: {total_formatted}"
    await message.answer(text)

@router.message(Command("заявка"))
async def start_form(message: Message, state: FSMContext):
    await state.set_state(Form.supplier)
    await message.answer("Введи поставщика:")

@router.message(Form.supplier)
async def step_supplier(message: Message, state: FSMContext):
    await state.update_data(supplier=message.text)
    await state.set_state(Form.amount)
    await message.answer("Введи сумму:")

@router.message(Form.amount)
async def step_amount(message: Message, state: FSMContext):
    try:
        amount = float(message.text.replace(',', '.').replace(' ', ''))
    except:
        return await message.answer("Некорректная сумма. Введи число.")
    await state.update_data(amount=amount)
    await state.set_state(Form.description)
    await message.answer("Введи описание (если нужно):")

@router.message(Form.description)
async def step_description(message: Message, state: FSMContext):
    await state.update_data(description=message.text)
    await state.set_state(Form.agent_name)
    await message.answer("Имя агента:")

@router.message(Form.agent_name)
async def step_agent_name(message: Message, state: FSMContext):
    await state.update_data(agent_name=message.text)
    await state.set_state(Form.agent_phone)
    await message.answer("Номер агента:")

@router.message(Form.agent_phone)
async def step_agent_phone(message: Message, state: FSMContext):
    await state.update_data(agent_phone=message.text)
    await state.set_state(Form.delivery_date)
    await message.answer("Дата поставки (дд.мм.гггг):")

@router.message(Form.delivery_date)
async def step_delivery_date(message: Message, state: FSMContext):
    try:
        date_obj = datetime.strptime(message.text, "%d.%m.%Y").date()
        await state.update_data(delivery_date=date_obj.strftime("%Y-%m-%d"))
        await state.set_state(Form.admin_name)
        await message.answer("Имя админа:")
    except:
        await message.answer("Неверный формат даты. Введи ДД.ММ.ГГГГ")

@router.message(Form.admin_name)
async def step_admin_name(message: Message, state: FSMContext):
    await state.update_data(admin_name=message.text)
    data = await state.get_data()
    user = message.from_user

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO requests (user_id, username, supplier, amount, description, agent_name, agent_phone, delivery_date, admin_name)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        user.id, user.full_name, data['supplier'], data['amount'], data.get('description', ''),
        data['agent_name'], data['agent_phone'], data['delivery_date'], data['admin_name']
    ))
    conn.commit()
    conn.close()

    await state.clear()
    await message.answer("✅ Заявка сохранена.")
   
    # Формируем сообщение для группы
    message_text = (
        f"📦 Новая заявка от {user.full_name}:\n\n"
        f"Поставщик: {data['supplier']}\n"
        f"Сумма: {data['amount']}\n"
    )
   
    if data.get('description'):
        message_text += f"Описание: {data['description']}\n"
       
    message_text += (
        f"Агент: {data['agent_name']}\n"
        f"Номер: {data['agent_phone']}\n"
        f"Дата поставки: {data['delivery_date']}\n"
        f"Админ: {data['admin_name']}\n"
    )
   
    await bot.send_message(GROUP_CHAT_ID, message_text)

@router.message(Command("заявки"))
async def list_requests(message: Message):
    parts = message.text.split()

    # Получаем дату, если указана
    if len(parts) > 1:
        try:
            target_date = datetime.strptime(parts[1], "%d.%m.%Y").date()
        except:
            return await message.answer("❌ Неверный формат даты. Используй: /заявки дд.мм.гггг")
    else:
        # Если дата не указана, берём сегодняшнюю
        target_date = datetime.now().date()

    # Определяем диапазон начала и конца дня
    date_start = datetime.combine(target_date, datetime.min.time())
    date_end = datetime.combine(target_date, datetime.max.time())

    # Запрос по дате СОЗДАНИЯ заявки (поле created_at)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        SELECT supplier, amount, description, agent_name, agent_phone, delivery_date, admin_name, username, created_at
        FROM requests
        WHERE created_at BETWEEN ? AND ?
    """, (date_start, date_end))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return await message.answer(f"📭 Нет заявок за {target_date.strftime('%d.%m.%Y')}.")

    total = sum(float(r[1]) for r in rows)
    total_formatted = "{:,.0f}".format(total).replace(",", ".")

    text = f"📦 Заявки за {target_date.strftime('%d.%m.%Y')}:\n\n"
    for i, r in enumerate(rows, 1):
        amount_formatted = "{:,.0f}".format(float(r[1])).replace(",", ".")
        text += (
            f"{i}) Поставщик: {r[0]}\n"
            f"Сумма: {amount_formatted}\n"
        )
        if r[2]:  # Добавляем описание, если оно есть
            text += f"Описание: {r[2]}\n"
        text += (
            f"Агент: {r[3]}\n"
            f"Номер: {r[4]}\n"
            f"Дата поставки: {datetime.strptime(r[5], '%Y-%m-%d').strftime('%d.%m.%Y')}\n"
            f"Админ: {r[6]}\n"
            f"От кого: {r[7]}\n\n"
        )

    text += f"💰 Общая сумма заявок: {total_formatted}"
    await message.answer(text)

@router.message(Command("экспорт"))
async def export_requests(message: Message):
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = generate_excel_by_date(date_str)
    if filename:
        await message.answer_document(FSInputFile(filename))
        os.remove(filename)
    else:
        await message.answer("Нет заявок для экспорта.")

async def scheduler():
    while True:
        now = datetime.now()
        if now.time().hour == 19 and now.time().minute == 00:
            date_str = now.strftime("%Y-%m-%d")
            filename = generate_excel_by_date(date_str)
            if filename:
                await bot.send_document(GROUP_CHAT_ID, FSInputFile(filename), caption=f"📄 Заявки за {date_str}")
                os.remove(filename)
        await asyncio.sleep(60)

async def main():
    init_db()
    dp.include_router(router)
    asyncio.create_task(scheduler())
    await dp.start_polling(bot, polling_timeout=30)

if __name__ == '__main__':
    asyncio.run(main())
